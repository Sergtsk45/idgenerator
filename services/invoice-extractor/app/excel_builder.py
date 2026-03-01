"""
app/excel_builder.py
Формирование Excel-файла из данных счёта.
Лист «Счет»: шапка → таблица товаров → итоги.
Лист «Метаданные»: технические поля для отладки.
"""
from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Цветовая палитра ─────────────────────────────────────────
C_HEADER_BG  = "1F4E79"  # тёмно-синий
C_HEADER_FG  = "FFFFFF"
C_SECTION_BG = "D6E4F0"  # светло-голубой
C_ALT_ROW    = "EBF5FB"  # очень светлый голубой
C_TOTAL_BG   = "FFF2CC"  # жёлтый
C_WARN_BG    = "FFE0E0"  # розовый

NUM_FMT  = '#,##0.00'
QTY_FMT  = '#,##0.###'


def build_excel(data: dict, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Счет"

    row = 1
    row = _write_header_block(ws, data, row)
    row += 1
    row = _write_items_table(ws, data.get("items", []), row)
    row += 1
    row = _write_totals_block(ws, data.get("totals", {}), row)
    row += 1
    row = _write_warnings_block(ws, data.get("warnings", []), row)

    _autofit_columns(ws)

    # Лист «Метаданные»
    ws2 = wb.create_sheet("Метаданные")
    _write_metadata_sheet(ws2, data)

    wb.save(output_path)


# ── Шапка счёта ───────────────────────────────────────────────

def _write_header_block(ws, data: dict, start_row: int) -> int:
    r = start_row
    _section_title(ws, r, "РЕКВИЗИТЫ СЧЁТА")
    r += 1

    fields = [
        ("Номер счёта",       data.get("invoice_number", "")),
        ("Дата счёта",        data.get("invoice_date", "")),
    ]
    sup = data.get("supplier", {})
    buy = data.get("buyer", {})
    bank = sup.get("bank", {})

    fields += [
        ("Поставщик",         sup.get("name", "")),
        ("ИНН поставщика",    sup.get("inn", "")),
        ("КПП поставщика",    sup.get("kpp", "")),
        ("Адрес поставщика",  sup.get("address", "")),
        ("Банк",              bank.get("name", "")),
        ("БИК",               bank.get("bik", "")),
        ("Расчётный счёт",    bank.get("account", "")),
        ("Корр. счёт",        bank.get("corr_account", "")),
        ("Покупатель",        buy.get("name", "")),
        ("ИНН покупателя",    buy.get("inn", "")),
        ("КПП покупателя",    buy.get("kpp", "")),
        ("Адрес покупателя",  buy.get("address", "")),
    ]

    for label, value in fields:
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=C_SECTION_BG)
        ws.cell(r, 2, str(value) if value else "")
        ws.cell(r, 2).alignment = Alignment(wrap_text=True)
        r += 1

    return r


# ── Таблица товаров ───────────────────────────────────────────

ITEMS_COLS = [
    ("№",           "line_no",       6,   None),
    ("Артикул",     "article",       15,  None),
    ("Наименование","name",          40,  None),
    ("Ед.",         "unit",          6,   None),
    ("Кол-во",      "qty",           9,   QTY_FMT),
    ("Цена",        "price",         13,  NUM_FMT),
    ("Скидка",      "discount",      11,  NUM_FMT),
    ("НДС %",       "vat_rate",      8,   None),
    ("Сумма НДС",   "vat_amount",    13,  NUM_FMT),
    ("Сумма с НДС", "amount_w_vat",  15,  NUM_FMT),
]


def _write_items_table(ws, items: list[dict], start_row: int) -> int:
    r = start_row
    _section_title(ws, r, "ПОЗИЦИИ СЧЁТА")
    r += 1

    # Заголовок таблицы
    for col_idx, (header, _, width, _) in enumerate(ITEMS_COLS, 1):
        cell = ws.cell(r, col_idx, header)
        cell.font      = Font(bold=True, color=C_HEADER_FG)
        cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[r].height = 30
    r += 1

    # Строки товаров
    for i, item in enumerate(items):
        bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=bg)
        for col_idx, (_, field, _, fmt) in enumerate(ITEMS_COLS, 1):
            value = item.get(field, "")
            cell = ws.cell(r, col_idx, value if value != "" else "")
            cell.fill   = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(
                horizontal="right" if fmt else "left",
                wrap_text=(field == "name"),
            )
            if fmt and value != "":
                cell.number_format = fmt
        r += 1

    return r


# ── Блок итогов ───────────────────────────────────────────────

def _write_totals_block(ws, totals: dict, start_row: int) -> int:
    r = start_row
    _section_title(ws, r, "ИТОГИ")
    r += 1

    rows = [
        ("Итого без НДС:",    totals.get("total_wo_vat", "")),
        ("НДС итого:",        totals.get("vat_total",    "")),
        ("ИТОГО К ОПЛАТЕ:",   totals.get("total_w_vat",  "")),
    ]
    for label, value in rows:
        cell_l = ws.cell(r, 1, label)
        cell_v = ws.cell(r, 2, value if value != "" else "")
        cell_l.font = Font(bold=True)
        cell_l.fill = PatternFill("solid", fgColor=C_TOTAL_BG)
        cell_v.fill = PatternFill("solid", fgColor=C_TOTAL_BG)
        cell_v.alignment = Alignment(horizontal="right")
        if value != "":
            cell_v.number_format = NUM_FMT
        r += 1

    return r


# ── Предупреждения ────────────────────────────────────────────

def _write_warnings_block(ws, warnings: list[str], start_row: int) -> int:
    if not warnings:
        return start_row
    r = start_row
    _section_title(ws, r, "⚠ ПРЕДУПРЕЖДЕНИЯ ВАЛИДАЦИИ", color=C_WARN_BG)
    r += 1
    for w in warnings:
        cell = ws.cell(r, 1, w)
        cell.fill = PatternFill("solid", fgColor=C_WARN_BG)
        ws.merge_cells(
            start_row=r, start_column=1,
            end_row=r, end_column=len(ITEMS_COLS)
        )
        r += 1
    return r


# ── Лист метаданных ───────────────────────────────────────────

def _write_metadata_sheet(ws, data: dict) -> None:
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    rows = [
        ("Страниц в PDF",     data.get("pages", "")),
        ("Тип документа",     data.get("document_type", "")),
        ("Сформировано",      datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Кол-во позиций",    len(data.get("items", []))),
    ]
    for r_idx, (k, v) in enumerate(rows, 1):
        ws.cell(r_idx, 1, k).font = Font(bold=True)
        ws.cell(r_idx, 2, str(v))


# ── Вспомогательные функции ───────────────────────────────────

def _section_title(ws, row: int, title: str, color: str = C_SECTION_BG) -> None:
    cell = ws.cell(row, 1, title)
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="left")
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=len(ITEMS_COLS)
    )


def _thin_border() -> Border:
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def _autofit_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                # Не учитываем объединённые ячейки и очень длинные строки
                if len(val) < 80:
                    max_len = max(max_len, len(val))
            except Exception:
                pass
        adjusted = min(max(max_len + 2, 8), 60)
        current = ws.column_dimensions[col_letter].width
        if current < adjusted:
            ws.column_dimensions[col_letter].width = adjusted
